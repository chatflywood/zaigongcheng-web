from fastapi import APIRouter, UploadFile, File, Query
from fastapi.responses import JSONResponse
import pandas as pd
from io import BytesIO
import math
import json
from datetime import datetime
from services.analysis import analyze, build_transfer_priority
from models import ZaigongRecord, AppConfig, get_db

router = APIRouter()

# 全年资本性支出目标（单位：万元）
YEARLY_TARGET = 503.0


def clean_nan(obj):
    """清理数据中的 NaN 和 Infinity 值"""
    if isinstance(obj, dict):
        return {k: clean_nan(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [clean_nan(item) for item in obj]
    elif isinstance(obj, float):
        if math.isnan(obj) or math.isinf(obj):
            return None
        return obj
    return obj


def build_dashboard_snapshot(record: ZaigongRecord) -> dict:
    """将数据库记录恢复为前端可直接使用的快照结构。"""
    raw_summary = json.loads(record.summary_data) if record.summary_data else []
    raw_metrics = json.loads(record.metrics_data) if record.metrics_data else {}
    detail_data = json.loads(record.detail_data) if record.detail_data else []
    four_class = json.loads(record.four_class_warnings) if record.four_class_warnings else {}

    detail_rows = [
        row for row in raw_summary
        if (row.get("工程管理员") or row.get("manager")) != "合计"
    ]
    pending_pressure = sum(1 for row in detail_rows if float(row.get("已下单待收货", 0) or row.get("pending", 0) or 0) > 30)
    total_rate = float(raw_metrics.get("total_rate", 0) or 0)
    deficit = float(raw_metrics.get("deficit", 0) or 0)

    return {
        "id": record.id,
        "uploaded_at": record.uploaded_at.isoformat(),
        "source_filename": record.source_filename,
        "file_date": record.file_date,
        "target_value": record.target_value,
        "dashboard": {
            "metrics": {
                "capital": raw_metrics.get("total_current", 0),
                "pending": raw_metrics.get("total_pending", 0),
                "monthSpend": raw_metrics.get("total_today_month", 0),
                "transfer": raw_metrics.get("total_transfer", 0),
                "rate": total_rate,
                "progress": raw_metrics.get("progress_ratio", 0),
                "deficit": deficit,
                "yearTarget": raw_metrics.get("year_target", record.target_value),
            },
            "summary": detail_rows,
            "detail": detail_data,
            "alerts": [
                {
                    "title": "待收货压力",
                    "value": f"{pending_pressure} 位管理员超过30万元",
                },
                {
                    "title": "转固率状态",
                    "value": "偏低" if total_rate < 0.6 else "正常",
                },
                {
                    "title": "目标差额",
                    "value": f"{deficit:.2f} 万元" if deficit > 0 else "已达成目标",
                },
            ],
        },
        "summary": raw_summary,
        "metrics": raw_metrics,
        "four_class_warnings": four_class,
    }


@router.post("/upload")
async def upload_excel(file: UploadFile = File(...), target: float = Query(503.0)):
    """
    上传 Excel 文件，返回分析结果，并保存到数据库
    """
    # 文件类型校验
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        return JSONResponse(status_code=400, content={"success": False, "message": "仅支持 Excel 文件（.xlsx / .xls）"})

    # 目标值校验
    if target <= 0:
        return JSONResponse(status_code=400, content={"success": False, "message": "目标值必须大于 0"})

    try:
        contents = await file.read()

        # 文件大小校验（20MB）
        if len(contents) > 20 * 1024 * 1024:
            return JSONResponse(status_code=400, content={"success": False, "message": "文件大小不能超过 20MB"})

        df = pd.read_excel(BytesIO(contents))

        # 执行分析
        result = analyze(df, year_target=target)

        # 清理 NaN 值
        cleaned_data = clean_nan(result["data"])

        # 从文件名提取日期（如 "在建工程明细总表(实时)(20260320).xlsx" -> "20260320"）
        import re
        # 匹配括号内的日期，如 (20260320) 或 (0320)
        date_match = re.search(r'\((\d{4,8})\)', file.filename)
        file_date = date_match.group(1) if date_match else None

        # 保存到数据库
        db = get_db()
        try:
            # 检查是否已存在相同文件名的记录，存在则删除
            existing = db.query(ZaigongRecord).filter(
                ZaigongRecord.source_filename == file.filename
            ).first()
            if existing:
                db.delete(existing)

            record = ZaigongRecord(
                uploaded_at=datetime.now(),
                source_filename=file.filename,
                file_date=file_date,
                summary_data=json.dumps(cleaned_data.get("summary", []), ensure_ascii=False),
                metrics_data=json.dumps(cleaned_data.get("metrics", {}), ensure_ascii=False),
                detail_data=json.dumps(cleaned_data.get("dashboard", {}).get("detail", []), ensure_ascii=False),
                four_class_warnings=json.dumps(cleaned_data.get("four_class_warnings", {}), ensure_ascii=False),
                target_value=target
            )
            db.add(record)
            db.commit()
        finally:
            db.close()

        # 自动推送（如果配置了 webhook 且开启了自动推送）
        try:
            db2 = get_db()
            webhook_row = db2.query(AppConfig).filter(AppConfig.key == "wework_webhook_url").first()
            auto_row = db2.query(AppConfig).filter(AppConfig.key == "wework_auto_push").first()
            db2.close()
            if webhook_row and webhook_row.value and auto_row and auto_row.value == "true":
                import asyncio
                from services.notify import push_record
                from models import BudgetRecord
                db3 = get_db()
                new_record = db3.query(ZaigongRecord).order_by(ZaigongRecord.id.desc()).first()
                budget_record = db3.query(BudgetRecord).order_by(BudgetRecord.id.desc()).first()
                db3.close()
                if new_record:
                    snapshot = build_dashboard_snapshot(new_record)
                    budget_data = json.loads(budget_record.budget_data) if budget_record and budget_record.budget_data else None
                    asyncio.create_task(push_record(webhook_row.value, snapshot, budget_data))
        except Exception:
            pass  # 推送失败不影响上传结果

        return {
            "success": True,
            "message": "分析完成",
            "filename": file.filename,
            "rows": len(df),
            "data": cleaned_data
        }
    except ValueError as e:
        return JSONResponse(
            status_code=400,
            content={"success": False, "message": str(e)}
        )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"success": False, "message": str(e)}
        )


@router.get("/history")
async def get_history(limit: int = Query(10, ge=1, le=100)):
    """
    获取历史记录列表（按实际上传时间降序）
    """
    db = get_db()
    try:
        records = db.query(ZaigongRecord).order_by(
            ZaigongRecord.uploaded_at.desc()
        ).limit(limit).all()

        return {
            "success": True,
            "data": [
                {
                    "id": r.id,
                    "uploaded_at": r.uploaded_at.isoformat(),
                    "source_filename": r.source_filename,
                    "file_date": r.file_date,
                    "target_value": r.target_value,
                    "metrics": json.loads(r.metrics_data) if r.metrics_data else {}
                }
                for r in records
            ]
        }
    finally:
        db.close()


@router.get("/history/{record_id}")
async def get_history_snapshot(record_id: int):
    """
    获取指定历史记录的完整快照，并附带该记录上一版数据用于对比。
    """
    db = get_db()
    try:
        all_records = db.query(ZaigongRecord).order_by(
            ZaigongRecord.uploaded_at.desc()
        ).all()

        if not all_records:
            return JSONResponse(
                status_code=404,
                content={"success": False, "message": "暂无历史记录"}
            )

        current_index = next((index for index, item in enumerate(all_records) if item.id == record_id), None)
        if current_index is None:
            return JSONResponse(
                status_code=404,
                content={"success": False, "message": "记录不存在"}
            )

        current_record = all_records[current_index]
        previous_record = all_records[current_index + 1] if current_index + 1 < len(all_records) else None

        return {
            "success": True,
            "data": {
                "current": build_dashboard_snapshot(current_record),
                "previous": build_dashboard_snapshot(previous_record) if previous_record else None
            }
        }
    finally:
        db.close()


@router.get("/compare")
async def compare_with_previous():
    """
    获取最近一次数据，用于对比
    - latest: 数据库中最近一次上传的记录
    - previous: 数据库中倒数第二次上传的记录（按实际上传时间 uploaded_at 排序）
    """
    db = get_db()
    try:
        # 按实际上传时间降序排序，获取所有记录
        all_records = db.query(ZaigongRecord).order_by(
            ZaigongRecord.uploaded_at.desc()
        ).all()

        if not all_records:
            return {"success": True, "data": None}

        # 获取最新的记录（最近一次上传）
        latest = all_records[0]

        # 获取上一条记录（按实际上传时间，排除同一条记录）
        previous = all_records[1] if len(all_records) > 1 else None

        result = {
            "latest": {
                "id": latest.id,
                "uploaded_at": latest.uploaded_at.isoformat(),
                "source_filename": latest.source_filename,
                "file_date": latest.file_date,
                "summary": json.loads(latest.summary_data) if latest.summary_data else [],
                "metrics": json.loads(latest.metrics_data) if latest.metrics_data else {},
                "four_class_warnings": json.loads(latest.four_class_warnings) if latest.four_class_warnings else {},
                "target_value": latest.target_value
            },
            "previous": None
        }

        if previous:
            result["previous"] = {
                "id": previous.id,
                "uploaded_at": previous.uploaded_at.isoformat(),
                "source_filename": previous.source_filename,
                "file_date": previous.file_date,
                "summary": json.loads(previous.summary_data) if previous.summary_data else [],
                "metrics": json.loads(previous.metrics_data) if previous.metrics_data else {},
                "four_class_warnings": json.loads(previous.four_class_warnings) if previous.four_class_warnings else {},
                "target_value": previous.target_value
            }

        return {"success": True, "data": result}
    finally:
        db.close()


@router.get("/manager-details")
async def get_manager_details(record_id: int = Query(...), manager: str = Query(...)):
    """
    获取指定记录中某工程管理员的明细数据
    """
    db = get_db()
    try:
        record = db.query(ZaigongRecord).filter(ZaigongRecord.id == record_id).first()
        if not record:
            return JSONResponse(
                status_code=404,
                content={"success": False, "message": "记录不存在"}
            )

        detail_data = json.loads(record.detail_data) if record.detail_data else []
        # 筛选指定管理员的明细
        filtered = [d for d in detail_data if d.get("工程管理员") == manager]

        return {
            "success": True,
            "data": {
                "manager": manager,
                "details": filtered
            }
        }
    finally:
        db.close()


@router.get("/metrics")
async def get_metrics():
    """
    获取核心指标
    """
    return {
        "success": True,
        "data": {
            "year_target": YEARLY_TARGET,
        }
    }


@router.get("/transfer-priority/{record_id}")
async def get_transfer_priority(record_id: int):
    """
    计算指定记录的转固推进优先级清单。
    基于已存储的 summary_data / detail_data / four_class_warnings 实时计算，
    返回各管理员按优先级排序的待转固项目清单及转固率提升模拟。
    """
    db = get_db()
    try:
        record = db.query(ZaigongRecord).filter(ZaigongRecord.id == record_id).first()
        if not record:
            return JSONResponse(status_code=404, content={"success": False, "message": "记录不存在"})

        summary_data   = json.loads(record.summary_data)   if record.summary_data   else []
        detail_data    = json.loads(record.detail_data)    if record.detail_data    else []
        four_class_data = json.loads(record.four_class_warnings) if record.four_class_warnings else {}

        priority_list = build_transfer_priority(summary_data, detail_data, four_class_data)
        return {"success": True, "data": priority_list}
    finally:
        db.close()


@router.get("/transfer-priority/{record_id}/export")
async def export_transfer_priority(record_id: int, target_rate: float = Query(None, ge=0.01, le=1.0)):
    """
    导出转固推进清单 Excel。
    target_rate: 可选，0~1 之间的小数，传入时在表格中标注"需完成"项目。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    db = get_db()
    try:
        record = db.query(ZaigongRecord).filter(ZaigongRecord.id == record_id).first()
        if not record:
            return JSONResponse(status_code=404, content={"success": False, "message": "记录不存在"})

        summary_data    = json.loads(record.summary_data)    if record.summary_data    else []
        detail_data     = json.loads(record.detail_data)     if record.detail_data     else []
        four_class_data = json.loads(record.four_class_warnings) if record.four_class_warnings else {}

        priority_list = build_transfer_priority(summary_data, detail_data, four_class_data)
        if not priority_list:
            return JSONResponse(status_code=404, content={"success": False, "message": "无待转固项目数据"})

        # ── 若传入目标转固率，计算各项目"是否需完成" ──
        def compute_needed(manager_group):
            if target_rate is None:
                return [None] * len(manager_group["projects"])
            if manager_group["current_rate"] >= target_rate:
                return [False] * len(manager_group["projects"])
            projects = manager_group["projects"]
            cutoff = next((i for i, p in enumerate(projects) if p["累计后转固率"] >= target_rate), -1)
            return [True if cutoff == -1 else i <= cutoff for i in range(len(projects))]

        # ── 全局统计 ──
        total_projects = sum(len(m["projects"]) for m in priority_list)
        file_date = record.file_date or ""

        # ── 颜色常量 ──
        C_NAVY    = "1B2A4A"
        C_BLUE    = "2E5F9E"
        C_WHITE   = "FFFFFF"
        C_MGR_BG  = "EEF3FB"   # 管理员组头背景（淡蓝）
        C_MGR_FG  = "1F3864"   # 管理员组头文字
        C_HEAD_BG = "F2F2F2"   # 列标题背景
        C_NEED_BG = "FFFDE7"   # 需完成行背景（淡黄）
        C_OPT_BG  = "FAFAFA"   # 可选行背景
        C_OVER    = "C00000"   # 已逾期红
        C_WARN    = "ED7D31"   # 即将到期橙
        C_GREEN   = "375623"   # 完成后转固率绿

        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def cell_style(ws, row, col, value, *, bold=False, size=9, color="000000",
                       bg=None, align="left", wrap=False, num_fmt=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font = Font(name="微软雅黑", size=size, bold=bold, color=color)
            if bg:
                c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
            c.border = border
            if num_fmt:
                c.number_format = num_fmt
            return c

        wb = Workbook()
        ws = wb.active
        ws.title = "转固推进清单"

        # ── 第1行：总标题 ──
        COLS = 9
        ws.merge_cells(f"A1:{get_column_letter(COLS)}1")
        ws["A1"].value = "中国电信股份有限公司仙桃分公司  在建工程转固推进清单"
        ws["A1"].font = Font(name="微软雅黑", size=15, bold=True, color=C_WHITE)
        ws["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 38

        # ── 第2行：副标题 ──
        ws.merge_cells(f"A2:{get_column_letter(COLS)}2")
        sub = f"数据日期：{file_date}  |  共 {len(priority_list)} 位管理员  |  {total_projects} 个项目待转固"
        if target_rate:
            sub += f"  |  测算目标转固率：{target_rate * 100:.1f}%"
        ws["A2"].value = sub
        ws["A2"].font = Font(name="微软雅黑", size=9, color="D9E8F5")
        ws["A2"].fill = PatternFill("solid", fgColor=C_BLUE)
        ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20

        # ── 第3行：图例 ──
        legend_row = 3
        ws.merge_cells(f"A{legend_row}:C{legend_row}")
        ws[f"A{legend_row}"].value = "已逾期：转固相关四类工程已触发"
        ws[f"A{legend_row}"].font = Font(name="微软雅黑", size=9, bold=True, color="7B0000")
        ws[f"A{legend_row}"].fill = PatternFill("solid", fgColor="FFD7D7")
        ws[f"A{legend_row}"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells(f"D{legend_row}:F{legend_row}")
        ws[f"D{legend_row}"].value = "即将到期：距截止日期不足60天"
        ws[f"D{legend_row}"].font = Font(name="微软雅黑", size=9, bold=True, color="7B3F00")
        ws[f"D{legend_row}"].fill = PatternFill("solid", fgColor="FFF3C8")
        ws[f"D{legend_row}"].alignment = Alignment(horizontal="center", vertical="center")

        if target_rate:
            ws.merge_cells(f"G{legend_row}:{get_column_letter(COLS)}{legend_row}")
            ws[f"G{legend_row}"].value = f"★ 需完成：完成该项目后累计转固率可达目标 {target_rate * 100:.1f}%"
            ws[f"G{legend_row}"].font = Font(name="微软雅黑", size=9, bold=True, color="7B5200")
            ws[f"G{legend_row}"].fill = PatternFill("solid", fgColor="FFFDE7")
            ws[f"G{legend_row}"].alignment = Alignment(horizontal="center", vertical="center")
        else:
            ws.merge_cells(f"G{legend_row}:{get_column_letter(COLS)}{legend_row}")
            ws[f"G{legend_row}"].value = "项目按转固贡献率从高到低排序"
            ws[f"G{legend_row}"].font = Font(name="微软雅黑", size=9, italic=True, color="595959")
            ws[f"G{legend_row}"].fill = PatternFill("solid", fgColor="F0F0F0")
            ws[f"G{legend_row}"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[legend_row].height = 20

        # ── 列宽 ──
        col_widths = [5, 46, 14, 12, 10, 14, 16, 14, 36]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── 列标题（每个管理员块前复用，此处写一个通用函数）──
        COL_HEADERS = ["序", "工程名称", "在建余额(万)", "转固贡献率", "紧迫度",
                       "完成后转固率", "累计提升(pct)", "任务标记" if target_rate else "完成建议", "紧迫说明"]

        current_row = 4

        for mgr in priority_list:
            needed_flags = compute_needed(mgr)
            current_rate = mgr["current_rate"]
            projects     = mgr["projects"]

            # ── 管理员组头 ──
            ws.merge_cells(f"A{current_row}:{get_column_letter(COLS)}{current_row}")
            mgr_label = f"▌ {mgr['manager']}    当前转固率：{current_rate * 100:.1f}%    待转固余额：{mgr['total_balance']:,.2f} 万元"
            if target_rate and current_rate < target_rate:
                needed_cnt = sum(1 for f in needed_flags if f is True)
                mgr_label += f"    目标 {target_rate*100:.1f}% → 需完成 {needed_cnt} 个项目"
            elif target_rate and current_rate >= target_rate:
                mgr_label += f"    ✓ 已达目标 {target_rate*100:.1f}%"
            c = ws[f"A{current_row}"]
            c.value = mgr_label
            c.font = Font(name="微软雅黑", size=10, bold=True, color=C_MGR_FG)
            c.fill = PatternFill("solid", fgColor=C_MGR_BG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border
            ws.row_dimensions[current_row].height = 22
            current_row += 1

            # ── 列标题行 ──
            for col, hdr in enumerate(COL_HEADERS, 1):
                c = ws.cell(row=current_row, column=col, value=hdr)
                c.font = Font(name="微软雅黑", size=9, bold=True, color="595959")
                c.fill = PatternFill("solid", fgColor=C_HEAD_BG)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = border
            ws.row_dimensions[current_row].height = 20
            current_row += 1

            # ── 项目明细行 ──
            for idx, (proj, needed) in enumerate(zip(projects, needed_flags), 1):
                urgency   = proj.get("紧迫度", "正常")
                after_rate = proj.get("累计后转固率", current_rate)
                delta_pct  = (after_rate - current_rate) * 100

                # 行背景
                if needed is True:
                    row_bg = C_NEED_BG
                elif urgency == "已逾期":
                    row_bg = "FFEDED"
                elif urgency == "即将到期":
                    row_bg = "FFF9E6"
                else:
                    row_bg = C_OPT_BG

                # 任务标记列
                if needed is True:
                    task_mark = "★ 需完成"
                elif needed is False:
                    task_mark = "可选缓冲"
                else:
                    # 无目标时，用紧迫度作为建议
                    task_mark = {"已逾期": "立即推进", "即将到期": "尽快推进", "正常": "按序推进"}.get(urgency, "")

                # 紧迫说明（取第一条）
                hint = ""
                for u in proj.get("urgency_detail", []):
                    hint = f"{u['type']}·{u['daysLabel']}"
                    if u.get("deadline"):
                        hint += f"·截止{u['deadline']}"
                    break

                row_data = [
                    idx,
                    proj.get("工程名称", ""),
                    proj.get("在建余额", 0),
                    proj.get("转固贡献率", 0),
                    urgency,
                    after_rate,
                    round(delta_pct, 1),
                    task_mark,
                    hint,
                ]

                for col, val in enumerate(row_data, 1):
                    c = ws.cell(row=current_row, column=col, value=val)
                    c.fill = PatternFill("solid", fgColor=row_bg)
                    c.alignment = Alignment(
                        horizontal="center" if col not in (2, 9) else "left",
                        vertical="center", wrap_text=(col == 9)
                    )
                    c.border = border

                    # 字体与特殊着色
                    fg = "000000"
                    bold = False
                    num_fmt = None

                    if col == 3:   # 在建余额
                        num_fmt = '#,##0.00'
                    elif col == 4: # 转固贡献率
                        val_pct = val * 100
                        c.value = val_pct / 100
                        num_fmt = '0.0%'
                    elif col == 5: # 紧迫度
                        if urgency == "已逾期":   fg = C_OVER
                        elif urgency == "即将到期": fg = C_WARN
                        else:                    fg = "595959"
                        bold = True
                    elif col == 6: # 完成后转固率
                        c.value = val
                        num_fmt = '0.0%'
                        fg = C_GREEN
                        bold = True
                    elif col == 7: # 累计提升
                        num_fmt = '+0.0;-0.0;0.0'
                        fg = "1A56A4"
                    elif col == 8: # 任务标记
                        if needed is True: fg, bold = "7B5200", True
                        elif needed is False: fg = "AAAAAA"

                    c.font = Font(name="微软雅黑", size=9, bold=bold, color=fg)
                    if num_fmt:
                        c.number_format = num_fmt

                ws.row_dimensions[current_row].height = 36
                current_row += 1

            # ── 管理员小计行 ──
            total_balance = mgr["total_balance"]
            ws.merge_cells(f"A{current_row}:B{current_row}")
            c = ws[f"A{current_row}"]
            c.value = f"{mgr['manager']} 合计"
            c.font = Font(name="微软雅黑", size=9, bold=True, color=C_MGR_FG)
            c.fill = PatternFill("solid", fgColor=C_MGR_BG)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = border

            c3 = ws.cell(row=current_row, column=3, value=total_balance)
            c3.font = Font(name="微软雅黑", size=9, bold=True, color=C_MGR_FG)
            c3.fill = PatternFill("solid", fgColor=C_MGR_BG)
            c3.alignment = Alignment(horizontal="center", vertical="center")
            c3.border = border
            c3.number_format = '#,##0.00'

            for col in range(4, COLS + 1):
                c = ws.cell(row=current_row, column=col, value="")
                c.fill = PatternFill("solid", fgColor=C_MGR_BG)
                c.border = border
            ws.row_dimensions[current_row].height = 18
            current_row += 1

            # 管理员之间空一行
            current_row += 1

        ws.freeze_panes = "A5"

        # ── Sheet 2: 说明 ──
        ws2 = wb.create_sheet("说明")
        ws2.column_dimensions["A"].width = 18
        ws2.column_dimensions["B"].width = 68
        ws2.merge_cells("A1:B1")
        ws2["A1"].value = "转固推进清单说明"
        ws2["A1"].font = Font(name="微软雅黑", size=13, bold=True, color=C_WHITE)
        ws2["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 28

        notes = [
            ("转固率公式",    "转固率 = 1 - (在建工程期末余额 + 工程物资) / (本年累计资本性支出 + 在建工程年初数 + 工程物资年初数)"),
            ("在建余额",     "在建工程期末余额（万元），减至 0 即代表该项目完成转固。"),
            ("转固贡献率",   "该项目完成转固后，管理员转固率的直接提升幅度（%）。"),
            ("完成后转固率", "按优先级依次完成本项目及之前所有项目后，管理员转固率的模拟值。"),
            ("项目排序规则", "① 四类预警已逾期 → ② 四类预警即将到期 → ③ 在建余额从大到小（余额越大，贡献越高）。"),
            ("任务标记",     "★需完成：达到目标转固率的最少必做项目。可选缓冲：完成需完成项后额外可推进的项目。"),
            ("数据来源",     "基于最近一次上传的在建工程明细总表快照，每次上传新数据后自动更新。"),
        ]
        for i, (k, v) in enumerate(notes, 2):
            bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
            for col, text in ((1, k), (2, v)):
                c = ws2.cell(row=i, column=col, value=text)
                c.font = Font(name="微软雅黑", size=10, bold=(col == 1))
                c.fill = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws2.row_dimensions[i].height = 32

        # ── 输出 ──
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse
        from urllib.parse import quote
        suffix = f"_{int(target_rate * 100)}pct目标" if target_rate else ""
        filename = f"转固推进清单_{file_date}{suffix}.xlsx"
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
        )
    finally:
        db.close()


@router.get("/four-class-warnings/{record_id}")
async def get_four_class_warnings(record_id: int):
    """
    获取指定记录的的四类工程预警数据
    """
    db = get_db()
    try:
        record = db.query(ZaigongRecord).filter(ZaigongRecord.id == record_id).first()
        if not record:
            return JSONResponse(
                status_code=404,
                content={"success": False, "message": "记录不存在"}
            )

        four_class = json.loads(record.four_class_warnings) if record.four_class_warnings else None
        return {
            "success": True,
            "data": four_class
        }
    finally:
        db.close()


@router.get("/four-class-warnings/{record_id}/export")
async def export_four_class_excel(record_id: int):
    """
    导出四类工程预警Excel文件（按Agent规范格式）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    db = get_db()
    try:
        record = db.query(ZaigongRecord).filter(ZaigongRecord.id == record_id).first()
        if not record:
            return JSONResponse(
                status_code=404,
                content={"success": False, "message": "记录不存在"}
            )

        four_class = json.loads(record.four_class_warnings) if record.four_class_warnings else None
        if not four_class or not four_class.get("items"):
            return JSONResponse(
                status_code=404,
                content={"success": False, "message": "无预警数据"}
            )

        wb = Workbook()

        # ===== Sheet1: 四类工程预警清单 =====
        ws1 = wb.active
        ws1.title = "四类工程预警清单"

        C_NAVY = "1B2A4A"
        C_BLUE = "2E5F9E"
        C_WHITE = "FFFFFF"

        # 类型颜色（分组标题背景/字色）
        TYPE_COLORS = {
            "列账不及时": ("DCE6F1", "1F497D"),
            "预转固不及时": ("FEF3E8", "7B3F00"),
            "关闭不及时": ("FFF0E5", "843C0C"),
            "长期挂账": ("DDE8F0", "244062"),
        }

        # 第1行：标题
        ws1.merge_cells("A1:M1")
        ws1["A1"].value = '中国电信股份有限公司仙桃分公司  "四类工程"预警清单'
        ws1["A1"].font = Font(name="微软雅黑", size=15, bold=True, color=C_WHITE)
        ws1["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
        ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[1].height = 38

        # 第2行：副标题
        today_str = four_class.get("analysis_date", "")
        total = four_class.get("total", 0)
        hit = four_class.get("hit_count", 0)
        warn = four_class.get("warn_count", 0)
        ws1.merge_cells("A2:M2")
        ws1["A2"].value = f"数据截止：{today_str}  |  排除局房类  |  预警窗口60天  |  共{total}项（已触发{hit}项/预警{warn}项）"
        ws1["A2"].font = Font(name="微软雅黑", size=9, color="D9E8F5")
        ws1["A2"].fill = PatternFill("solid", fgColor=C_BLUE)
        ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[2].height = 20

        # 第3行：图例三段
        ws1.merge_cells("A3:D3")
        ws1["A3"].value = "🔴 已触发：指标已超红线，须立即处置"
        ws1["A3"].font = Font(name="微软雅黑", size=9, bold=True, color="7B0000")
        ws1["A3"].fill = PatternFill("solid", fgColor="FFD7D7")
        ws1["A3"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("E3:H3")
        ws1["E3"].value = "🟡 预警：距红线不足60天，需重点跟进"
        ws1["E3"].font = Font(name="微软雅黑", size=9, bold=True, color="7B5200")
        ws1["E3"].fill = PatternFill("solid", fgColor="FFF3C8")
        ws1["E3"].alignment = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("I3:M3")
        ws1["I3"].value = "局房及基础设施、保密项目不纳入四类工程考核范围"
        ws1["I3"].font = Font(name="微软雅黑", size=9, italic=True, color="595959")
        ws1["I3"].fill = PatternFill("solid", fgColor="F0F0F0")
        ws1["I3"].alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[3].height = 20

        # 第4行：表头
        headers = ["编号", "状态", "四类工程类型", "工程编码", "工程名称", "一级专业", "验收类型", "工程管理员", "关键日期", "关键日期说明", "截止日期", "工程状态", "逾期/剩余天数", "处置建议"]
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=4, column=col)
            cell.value = header
            cell.font = Font(name="微软雅黑", size=10, bold=True, color=C_WHITE)
            cell.fill = PatternFill("solid", fgColor=C_NAVY)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[4].height = 28

        # 列宽
        col_widths = [5, 12, 13, 18, 44, 10, 8, 9, 11, 12, 11, 12, 40]
        for col, width in enumerate(col_widths, 1):
            ws1.column_dimensions[get_column_letter(col)].width = width

        # 数据行
        items = four_class.get("items", [])
        current_row = 5

        # 按类型分组
        types_order = ["列账不及时", "预转固不及时", "关闭不及时", "长期挂账"]
        TYPE_STANDARDS = {
            "列账不及时": "初验批复后累计收货占比 < 85%",
            "预转固不及时": "初验批复后60天内未完成预转固",
            "关闭不及时": "一次验收终验后150天/两次验收终验后90天内未正式转固",
            "长期挂账": "实际工期超建议工期2倍",
        }

        for wtype in types_order:
            type_items = [it for it in items if it.get("type") == it.get("type") == wtype]
            if not type_items:
                continue

            triggered_count = sum(1 for it in type_items if it.get("status", "").startswith("已触发"))
            warning_count = sum(1 for it in type_items if it.get("status") == "预警")

            # 分组标题行
            bg_color, text_color = TYPE_COLORS.get(wtype, ("F0F0F0", "595959"))
            ws1.merge_cells(f"A{current_row}:M{current_row}")
            cell = ws1[f"A{current_row}"]
            cell.value = f"▌ {wtype}　{TYPE_STANDARDS.get(wtype, '')}　已触发：{triggered_count}项　预警：{warning_count}项"
            cell.font = Font(name="微软雅黑", size=10, bold=True, color=text_color)
            cell.fill = PatternFill("solid", fgColor=bg_color)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            ws1.row_dimensions[current_row].height = 22
            current_row += 1

            # 数据行
            for item in type_items:
                status = item.get("status", "")
                is_triggered = status.startswith("已触发")

                # 行背景色
                if is_triggered:
                    row_bg = "FFEDED"
                    status_color = "C00000"
                else:
                    row_bg = "FFF9E6"
                    status_color = "7B5200"

                row_data = [
                    item.get("id", ""),
                    status,
                    item.get("type", ""),
                    item.get("code", ""),
                    item.get("name", ""),
                    item.get("major", ""),
                    item.get("acceptType", ""),
                    item.get("manager", ""),
                    item.get("keyDate", ""),
                    item.get("keyDateLabel", ""),
                    item.get("deadline", ""),
                    item.get("projectStatus", ""),
                    item.get("daysLabel", ""),
                    item.get("suggestion", ""),
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws1.cell(row=current_row, column=col)
                    cell.value = value
                    cell.font = Font(name="微软雅黑", size=9, color=text_color if col == 3 else "000000")
                    cell.fill = PatternFill("solid", fgColor=row_bg)
                    cell.alignment = Alignment(horizontal="left" if col in [5, 13] else "center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=Side(style="thin", color="BFBFBF"),
                        right=Side(style="thin", color="BFBFBF"),
                        top=Side(style="thin", color="BFBFBF"),
                        bottom=Side(style="thin", color="BFBFBF"),
                    )

                    # B列状态颜色
                    if col == 2:
                        cell.font = Font(name="微软雅黑", size=9, bold=True, color=status_color)

                    # L列天数颜色
                    if col == 12:
                        days_text = str(value)
                        if "逾期" in days_text or "超期" in days_text:
                            cell.font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
                        elif "剩余" in days_text:
                            import re
                            match = re.search(r'(\d+)', days_text)
                            if match:
                                num = int(match.group(1))
                                if num <= 15:
                                    cell.font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
                                elif num <= 30:
                                    cell.font = Font(name="微软雅黑", size=9, bold=True, color="ED7D31")
                                else:
                                    cell.font = Font(name="微软雅黑", size=9, color="375623")

                ws1.row_dimensions[current_row].height = 40
                current_row += 1

        # 冻结窗格
        ws1.freeze_panes = "A5"

        # ===== Sheet2: 四类工程定义说明 =====
        ws2 = wb.create_sheet(title="四类工程定义说明")
        ws2.column_dimensions["A"].width = 16
        ws2.column_dimensions["B"].width = 72

        ws2.merge_cells("A1:B1")
        ws2["A1"].value = "四类工程定义说明"
        ws2["A1"].font = Font(name="微软雅黑", size=14, bold=True, color=C_WHITE)
        ws2["A1"].fill = PatternFill("solid", fgColor=C_NAVY)
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 30

        def_content = [
            ("① 列账不及时", "初验批复后，累计收货金额÷累计订单金额<85%则触发；85%~90%预警。"),
            ("② 预转固不及时", "初验批复起60天内须完成预转固（取预转固日期，无则取决算转固日期）；截止前60天内未完成则预警。"),
            ("③ 关闭不及时", "一次验收终验批复后150天、两次验收90天内须完成正式转固（工程关闭）；截止前60天内预警。优先取系统字段「应关闭日期」。"),
            ("④ 长期挂账", "实际工期超建议工期2倍（有线接入9月/无线1年/省内1.5年/一干2年）；系统字段「长期挂账建议关闭日期」直接使用，距此日期60天内预警。"),
            ("排除范围", "局房及基础设施类工程、保密项目不纳入四类工程考核范围，在过滤阶段排除。"),
            ("预警窗口", "本清单预警窗口设定为60天，即距截止日期不足60天纳入预警提示。"),
            ("数据说明", "数据来源：在建工程明细总表（实时）。分析对象：排除局房类和已关闭工程后的全部在建工程。"),
        ]

        for i, (col_a, col_b) in enumerate(def_content, 2):
            bg = "EBF3FB" if i % 2 == 0 else "FFFFFF"
            ws2.cell(row=i, column=1).value = col_a
            ws2.cell(row=i, column=1).font = Font(name="微软雅黑", size=10, bold=True)
            ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor=bg)
            ws2.cell(row=i, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            ws2.cell(row=i, column=2).value = col_b
            ws2.cell(row=i, column=2).font = Font(name="微软雅黑", size=10)
            ws2.cell(row=i, column=2).fill = PatternFill("solid", fgColor=bg)
            ws2.cell(row=i, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            ws2.row_dimensions[i].height = 35

        # 保存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        from fastapi.responses import StreamingResponse
        from urllib.parse import quote
        filename = f"四类工程预警清单_{record.file_date or 'export'}.xlsx"
        encoded_filename = quote(filename)
        return StreamingResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )
    finally:
        db.close()
